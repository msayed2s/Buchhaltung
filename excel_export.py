from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from utils import fmt_date


def export_trips_excel(path, trips, year, company):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Fahrtkosten {year}"[:31]

    ws.append([f"Fahrtkostenaufstellung {year} — betriebliche Fahrten mit privatem PKW (Kilometerpauschale)"])
    ws.append([f"{company.get('name','')} · Inhaber: {company.get('owner','')} · "
               f"{company.get('addressLine1','')}, {company.get('addressLine2','')}"])
    ws.append(["Hinweis: Die angegebenen Kilometer sind die tatsächlich gefahrene Strecke (Hin- und Rückfahrt)."])
    ws.append([])
    header = ["Datum", "Anlass / Zweck der Fahrt", "Start", "Ziel",
              "Gefahrene km (hin + zurück)", "Satz €/km", "Betrag €"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    total_km = 0.0
    total_eur = 0.0
    rows = sorted(trips, key=lambda t: t.get("date") or "")
    for t in rows:
        km = float(t.get("km") or 0)
        rate = float(t.get("rate") or 0)
        amount = round(km * rate, 2)
        total_km += km
        total_eur += amount
        ws.append([fmt_date(t.get("date", "")), t.get("purpose", ""), t.get("origin", ""),
                   t.get("destination", ""), km, rate, amount])

    ws.append([])
    sum_row = ["Summe", "", "", "", total_km, "", round(total_eur, 2)]
    ws.append(sum_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    ws.append([])
    ws.append(["Hinweis: Ansatz als Betriebsausgabe (Nutzungseinlage) gem. Kilometerpauschale "
               "für betrieblich veranlasste Fahrten mit dem Privat-PKW."])

    widths = [12, 42, 28, 28, 24, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(path)
    return path
